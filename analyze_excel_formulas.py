import openpyxl
import json
import sys

def main():
    try:
        wb = openpyxl.load_workbook('IntermediateData.xlsx', data_only=False)
        ws = wb.active
        formulas = []
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                if cell.data_type == 'f':
                    formulas.append({
                        'cell': cell.coordinate,
                        'formula': cell.value,
                        'value': cell.value if cell.value else None
                    })
        print(f"Total formulas found: {len(formulas)}")
        # Print first 20 formulas
        for f in formulas[:20]:
            print(f"{f['cell']}: {f['formula']}")
        # Save all formulas to a JSON file for inspection
        with open('formulas.json', 'w', encoding='utf-8') as f:
            json.dump(formulas, f, ensure_ascii=False, indent=2)
        print("Formulas saved to formulas.json")
    except Exception as e:
        print(f"Error: {e}")
        sys.exit(1)

if __name__ == '__main__':
    main()