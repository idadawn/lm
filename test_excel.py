import openpyxl
import traceback

try:
    print("Loading workbook...")
    wb = openpyxl.load_workbook('IntermediateData.xlsx', data_only=False)
    print(f"Workbook loaded. Sheets: {wb.sheetnames}")

    ws = wb.active
    print(f"Active sheet: {ws.title}")
    print(f"Max row: {ws.max_row}, Max column: {ws.max_column}")

    # 检查前几行几列
    for row in ws.iter_rows(min_row=1, max_row=5, min_col=1, max_col=10):
        for cell in row:
            if cell.value:
                print(f"{cell.coordinate}: {cell.value[:50] if isinstance(cell.value, str) else cell.value}")

except Exception as e:
    print(f"Error: {e}")
    traceback.print_exc()